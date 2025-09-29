from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta, timezone
from jose import JWTError, jwt
import hashlib
import secrets
from fastapi_mail import FastMail, MessageSchema, ConnectionConfig, MessageType
from jinja2 import Template
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT and security setup
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

security = HTTPBearer()

# Email Configuration
conf = ConnectionConfig(
    MAIL_USERNAME=os.environ.get('MAIL_USERNAME'),
    MAIL_PASSWORD=os.environ.get('MAIL_PASSWORD'),
    MAIL_FROM=os.environ.get('MAIL_FROM'),
    MAIL_PORT=int(os.environ.get('MAIL_PORT', 587)),
    MAIL_SERVER=os.environ.get('MAIL_SERVER'),
    MAIL_FROM_NAME=os.environ.get('MAIL_FROM_NAME'),
    MAIL_STARTTLS=os.environ.get('MAIL_STARTTLS', 'true').lower() == 'true',
    MAIL_SSL_TLS=os.environ.get('MAIL_SSL_TLS', 'false').lower() == 'true',
    USE_CREDENTIALS=True,
    VALIDATE_CERTS=True
)

frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')

# Email Templates
EMAIL_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Dijital Dönüşüm - Soru Yanıtlama</title>
    <style>
        body { font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }
        .container { max-width: 600px; margin: 0 auto; background: white; border-radius: 8px; padding: 30px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { text-align: center; border-bottom: 2px solid #10b981; padding-bottom: 20px; margin-bottom: 30px; }
        .title { color: #1f2937; font-size: 24px; font-weight: bold; margin: 0; }
        .subtitle { color: #6b7280; font-size: 16px; margin-top: 5px; }
        .question-box { background: #f0fdfa; border-left: 4px solid #10b981; padding: 20px; margin: 20px 0; border-radius: 4px; }
        .question-title { font-weight: bold; color: #064e3b; margin-bottom: 10px; }
        .question-text { color: #374151; line-height: 1.6; }
        .button { display: inline-block; background: linear-gradient(135deg, #10b981, #3b82f6); color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; margin: 20px 0; }
        .button:hover { opacity: 0.9; }
        .footer { border-top: 1px solid #e5e7eb; padding-top: 20px; margin-top: 30px; text-align: center; color: #6b7280; font-size: 14px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1 class="title">Dijital Dönüşüm</h1>
            <p class="subtitle">Soru Yanıtlama Sistemi</p>
        </div>
        
        <p>Merhaba <strong>{{ employee_name }}</strong>,</p>
        
        <p>{{ month_year }} dönemi için size atanmış bir soru bulunmaktadır. Lütfen aşağıdaki soruyu inceleyip yanıtlayınız:</p>
        
        <div class="question-box">
            <div class="question-title">{{ question_category }} - {{ question_text }}</div>
            <div class="question-text">
                <strong>Önem/Gerekçe:</strong><br>
                {{ importance_reason }}
            </div>
        </div>
        
        <div style="text-align: center; margin: 30px 0;">
            <a href="{{ answer_link }}" class="button">Soruyu Yanıtla</a>
        </div>
        
        <p><strong>Not:</strong> Bu link sadece sizin için oluşturulmuştur ve tek kullanımlıktır.</p>
        
        <div class="footer">
            <p>Bu e-posta Dijital Dönüşüm sistemi tarafından otomatik olarak gönderilmiştir.</p>
            <p>Sorularınız için sistem yöneticinizle iletişime geçebilirsiniz.</p>
        </div>
    </div>
</body>
</html>
"""

async def send_question_email(employee_email: str, employee_name: str, question: dict, assignment_id: str, month_year: str):
    """Send question email to employee"""
    try:
        # Create email content
        answer_link = f"{frontend_url}/answer/{assignment_id}"
        
        template = Template(EMAIL_TEMPLATE)
        html_content = template.render(
            employee_name=employee_name,
            question_category=question.get('category', ''),
            question_text=question.get('question_text', ''),
            importance_reason=question.get('importance_reason', ''),
            answer_link=answer_link,
            month_year=month_year
        )
        
        # For demo purposes - simulate successful email sending
        print("[DEMO EMAIL] E-posta gönderildi:")
        print(f"Alıcı: {employee_email}")
        print(f"Konu: Dijital Dönüşüm - {month_year} Dönemi Soru Yanıtlama")
        print(f"İçerik: {employee_name} için {question.get('category', '')} sorusu")
        print(f"Link: {answer_link}")
        print("-" * 50)
        
        # Comment out real email sending for demo
        # message = MessageSchema(
        #     subject=f"Dijital Dönüşüm - {month_year} Dönemi Soru Yanıtlama",
        #     recipients=[employee_email],
        #     body=html_content,
        #     subtype=MessageType.html
        # )
        # 
        # fm = FastMail(conf)
        # await fm.send_message(message)
        
        return True  # Simulate successful sending
        
    except Exception as e:
        print(f"Email gönderim hatası: {str(e)}")
        return False

# AI Integration for Comment Generation
async def generate_ai_comment(question_text: str, category: str, period: str, table_data: Dict[str, str], table_rows: List[dict], monthly_comment: str = None, year: int = None, month: int = None) -> str:
    """Generate AI comment based on employee response"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        # Get the API key from environment
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        if not api_key:
            return "AI yorum oluşturma servisi şu anda kullanılamıyor."
        
        # Create AI chat instance
        chat = LlmChat(
            api_key=api_key,
            session_id=f"ai_comment_{str(uuid.uuid4())}",
            system_message="""Sen bir dijital dönüşüm uzmanısın. Çalışanların verdikleri yanıtları analiz edip yapıcı, profesyonel ve gelişim odaklı yorumlar yapıyorsun. 
            
            Görevin:
            1. Çalışanın verdiği yanıtı objektif olarak değerlendirmek
            2. Güçlü yönleri vurgulamak
            3. Gelişim alanları önermek
            4. Sayısal değerler için: trend analizi, büyüklük değerlendirmesi, benchmark karşılaştırması
            5. Kısa, net ve yapıcı olmak (maksimum 200 kelime)
            
            Sayısal değerler milyonlar, yüzdeler, adetler vb. herhangi bir formatda olabilir.
            Yanıtın Türkçe olmalı ve profesyonel bir ton kullanmalısın."""
        ).with_model("openai", "gpt-5")
        
        # Prepare table data text
        months_tr = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                     "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        month_name = months_tr[month - 1] if month and 1 <= month <= 12 else "Bilinmeyen Ay"
        
        data_text = f"\n{month_name} {year} Verileri:\n"
        if table_data and table_rows:
            row_dict = {row["id"]: row for row in table_rows}
            for row_id, value in table_data.items():
                if row_id in row_dict and value:
                    row = row_dict[row_id]
                    unit = f" {row['unit']}" if row.get('unit') else ""
                    data_text += f"- {row['name']}: {value}{unit}\n"
        
        comment_text = f"\nÇalışan Yorumu: {monthly_comment}" if monthly_comment else ""
        
        prompt = f"""
        Soru Kategorisi: {category}
        Soru: {question_text}
        Periyot: {period}{data_text}{comment_text}
        
        Bu aylık verileri analiz edip yapıcı bir AI yorumu oluştur. Değerlerin anlamı, trendler ve potansiyel iyileştirme alanları hakkında yorum yap.
        """
        
        # Create user message
        user_message = UserMessage(text=prompt)
        
        # Get AI response
        response = await chat.send_message(user_message)
        
        return response if response else "AI yorumu oluşturulamadı."
        
    except Exception as e:
        print(f"AI yorum oluşturma hatası: {str(e)}")
        return f"AI yorum oluşturulurken bir hata oluştu: {str(e)}"

# Create the main app without a prefix
app = FastAPI(title="Auth System API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Helper functions
def verify_password(plain_password, hashed_password):
    # Extract salt and hash
    parts = hashed_password.split('$')
    if len(parts) != 3:
        return False
    
    salt = parts[1]
    stored_hash = parts[2]
    
    # Hash the plain password with the same salt
    calculated_hash = hashlib.pbkdf2_hmac('sha256', plain_password.encode('utf-8'), salt.encode('utf-8'), 100000)
    calculated_hash_hex = calculated_hash.hex()
    
    return secrets.compare_digest(calculated_hash_hex, stored_hash)

def get_password_hash(password):
    # Generate a random salt
    salt = secrets.token_hex(16)
    
    # Hash the password with the salt
    password_hash = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 100000)
    password_hash_hex = password_hash.hex()
    
    # Return salt$hash format
    return f"pbkdf2_sha256${salt}${password_hash_hex}"

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Could not validate credentials",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="User not found",
                headers={"WWW-Authenticate": "Bearer"},
            )
        return User(**user)
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Could not validate credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=20)
    email: str
    password: str = Field(..., min_length=6)

class UserLogin(BaseModel):
    username: str
    password: str

class UserInDB(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    hashed_password: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class StatusCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class Employee(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str = Field(..., min_length=2, max_length=50)
    last_name: str = Field(..., min_length=2, max_length=50)
    phone: str = Field(..., min_length=10, max_length=15)
    email: Optional[str] = Field(None, min_length=5, max_length=100)
    department: str = Field(..., min_length=2, max_length=100)
    age: int = Field(..., ge=16, le=100)
    gender: str = Field(..., pattern="^(Erkek|Kadın|Diğer)$")
    hire_date: str  # YYYY-MM-DD format
    birth_date: str  # YYYY-MM-DD format
    salary: float = Field(..., ge=0)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeCreate(BaseModel):
    first_name: str = Field(..., min_length=2, max_length=50)
    last_name: str = Field(..., min_length=2, max_length=50)
    phone: str = Field(..., min_length=10, max_length=15)
    email: Optional[str] = Field(None, min_length=5, max_length=100)
    department: str = Field(..., min_length=2, max_length=100)
    age: int = Field(..., ge=16, le=100)
    gender: str = Field(..., pattern="^(Erkek|Kadın|Diğer)$")
    hire_date: str  # YYYY-MM-DD format
    birth_date: str  # YYYY-MM-DD format
    salary: float = Field(..., ge=0)

# Table Row Model for dynamic table structure
class TableRow(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str = Field(..., min_length=1, max_length=100)  # e.g., "Satış", "Pazarlama"
    unit: Optional[str] = Field(None, max_length=20)  # e.g., "adet", "TL", "%"
    order: int = Field(default=0)

class Question(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str = Field(..., min_length=2, max_length=100)
    question_text: str = Field(..., min_length=10, max_length=1000)
    importance_reason: str = Field(..., min_length=10, max_length=1000)
    expected_action: str = Field(..., min_length=10, max_length=1000)
    period: str = Field(..., pattern="^(Günlük|Haftalık|Aylık|Çeyreklik|Altı Aylık|Yıllık|İhtiyaç Halinde)$")
    chart_type: Optional[str] = Field(None, pattern="^(Sütun|Pasta|Çizgi|Alan|Daire|Bar|Trend)$")
    table_rows: List[TableRow] = Field(default_factory=list)  # Dynamic table rows (2-10 rows)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuestionCreate(BaseModel):
    category: str = Field(..., min_length=2, max_length=100)
    question_text: str = Field(..., min_length=10, max_length=1000)
    importance_reason: str = Field(..., min_length=10, max_length=1000)
    expected_action: str = Field(..., min_length=10, max_length=1000)
    period: str = Field(..., pattern="^(Günlük|Haftalık|Aylık|Çeyreklik|Altı Aylık|Yıllık|İhtiyaç Halinde)$")
    chart_type: Optional[str] = Field(None, pattern="^(Sütun|Pasta|Çizgi|Alan|Daire|Bar|Trend)$")
    table_rows: List[TableRow] = Field(default_factory=list)

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str = Field(..., min_length=2, max_length=100)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=100)

class Department(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str = Field(..., min_length=2, max_length=100)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DepartmentCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=100)

class QuestionAssignment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    question_id: str
    employee_id: str
    year: int
    month: int
    email_sent: bool = False
    response_received: bool = False
    assigned_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuestionResponse(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    assignment_id: str
    question_id: str
    employee_id: str
    response_text: str = Field(..., min_length=1, max_length=2000)
    year: int
    month: int
    submitted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuestionResponseCreate(BaseModel):
    assignment_id: str
    response_text: str = Field(..., min_length=1, max_length=2000)

class ShareQuestionsRequest(BaseModel):
    assignments: List[dict]  # [{"question_id": "...", "employee_id": "..."}]

# Response Data Models for Cevaplar Feature
# Table Response Model - Clean and Simple
class TableResponse(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    question_id: str
    employee_id: str
    year: int
    month: int
    # Table data: {"row_id": value, "row_id": value, ...}
    table_data: Dict[str, str] = Field(default_factory=dict)  # row_id -> value mapping
    monthly_comment: Optional[str] = Field(None, max_length=2000)  # Comment for this specific month
    ai_comment: Optional[str] = Field(None, max_length=3000)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TableResponseCreate(BaseModel):
    question_id: str
    employee_id: str  
    year: int
    month: int
    table_data: Dict[str, str] = Field(default_factory=dict)
    monthly_comment: Optional[str] = Field(None, max_length=2000)

class TableResponseUpdate(BaseModel):
    table_data: Dict[str, str] = Field(default_factory=dict)
    monthly_comment: Optional[str] = Field(None, max_length=2000)

# Auth Routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    # Check if user already exists
    existing_user = await db.users.find_one({"$or": [{"username": user_data.username}, {"email": user_data.email}]})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username or email already registered"
        )
    
    # Hash password
    hashed_password = get_password_hash(user_data.password)
    
    # Create user
    user_dict = {
        "id": str(uuid.uuid4()),
        "username": user_data.username,
        "email": user_data.email,
        "hashed_password": hashed_password,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_dict)
    
    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user_data.username}, expires_delta=access_token_expires
    )
    
    user = User(
        id=user_dict["id"],
        username=user_dict["username"],
        email=user_dict["email"],
        created_at=datetime.fromisoformat(user_dict["created_at"].replace('Z', '+00:00')) if isinstance(user_dict["created_at"], str) else user_dict["created_at"]
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    user = await db.users.find_one({"username": user_credentials.username})
    if not user or not verify_password(user_credentials.password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    
    user_obj = User(
        id=user["id"],
        username=user["username"],
        email=user["email"],
        created_at=datetime.fromisoformat(user["created_at"].replace('Z', '+00:00')) if isinstance(user["created_at"], str) else user["created_at"]
    )
    
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    return current_user

# Protected routes
@api_router.get("/analytics/insights/{question_id}")
async def get_advanced_insights(question_id: str, current_user: User = Depends(get_current_user)):
    """Get AI-powered advanced analytics insights for a specific question"""
    from datetime import datetime, timedelta
    import statistics
    
    # Get question data
    question = await db.questions.find_one({"id": question_id})
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    # Get all responses for this question
    responses = await db.table_responses.find({"question_id": question_id}).to_list(1000)
    
    if not responses:
        return {
            "question_id": question_id,
            "question_text": question.get("question_text", ""),
            "insights": {
                "data_trends": [],
                "predictions": [],
                "anomalies": [],
                "recommendations": ["Henüz yeterli veri bulunmuyor. Lütfen daha fazla yanıt toplayın."],
                "performance_score": 0,
                "confidence_level": "low"
            },
            "generated_at": datetime.now().isoformat()
        }
    
    # Organize data by periods and table rows
    period_data = {}
    table_rows = question.get("table_rows", [])
    
    for response in responses:
        period_key = f"{response.get('year', 0)}-{response.get('month', 0):02d}"
        if period_key not in period_data:
            period_data[period_key] = {}
        
        for row in table_rows:
            value = response.get("table_data", {}).get(row["id"], "0")
            try:
                numeric_value = float(value) if value else 0
                period_data[period_key][row["name"]] = numeric_value
            except (ValueError, TypeError):
                period_data[period_key][row["name"]] = 0
    
    # Calculate insights
    insights = {
        "data_trends": [],
        "predictions": [],
        "anomalies": [],
        "recommendations": [],
        "performance_score": 0,
        "confidence_level": "medium"
    }
    
    # Trend Analysis
    for row in table_rows:
        row_name = row["name"]
        values = []
        periods = []
        
        for period_key in sorted(period_data.keys()):
            if row_name in period_data[period_key]:
                values.append(period_data[period_key][row_name])
                periods.append(period_key)
        
        if len(values) >= 2:
            # Calculate trend
            if len(values) >= 3:
                recent_avg = statistics.mean(values[-3:])
                older_avg = statistics.mean(values[:-3]) if len(values) > 3 else values[0]
                trend_percentage = ((recent_avg - older_avg) / older_avg * 100) if older_avg != 0 else 0
            else:
                trend_percentage = ((values[-1] - values[0]) / values[0] * 100) if values[0] != 0 else 0
            
            trend_direction = "artış" if trend_percentage > 5 else "azalış" if trend_percentage < -5 else "stabil"
            
            insights["data_trends"].append({
                "metric": row_name,
                "direction": trend_direction,
                "percentage": round(trend_percentage, 2),
                "current_value": values[-1],
                "previous_value": values[-2] if len(values) >= 2 else 0,
                "confidence": "high" if len(values) >= 4 else "medium"
            })
            
            # Simple prediction (linear trend)
            if len(values) >= 3:
                # Simple linear regression for next period
                x_values = list(range(len(values)))
                slope = sum((x_values[i] - statistics.mean(x_values)) * (values[i] - statistics.mean(values)) 
                           for i in range(len(values))) / sum((x - statistics.mean(x_values))**2 for x in x_values)
                intercept = statistics.mean(values) - slope * statistics.mean(x_values)
                next_prediction = slope * len(values) + intercept
                
                insights["predictions"].append({
                    "metric": row_name,
                    "predicted_value": round(max(0, next_prediction), 2),
                    "confidence_interval": {
                        "min": round(max(0, next_prediction * 0.85), 2),
                        "max": round(next_prediction * 1.15, 2)
                    },
                    "period": "Gelecek dönem"
                })
            
            # Anomaly detection (simple outlier detection)
            if len(values) >= 4:
                mean_val = statistics.mean(values)
                std_dev = statistics.stdev(values) if len(values) > 1 else 0
                
                for i, value in enumerate(values):
                    if std_dev > 0 and abs(value - mean_val) > 2 * std_dev:
                        insights["anomalies"].append({
                            "metric": row_name,
                            "period": periods[i],
                            "value": value,
                            "expected_range": {
                                "min": round(mean_val - std_dev, 2),
                                "max": round(mean_val + std_dev, 2)
                            },
                            "severity": "high" if abs(value - mean_val) > 3 * std_dev else "medium",
                            "description": f"{row_name} değeri normal aralığın dışında ({value} vs beklenen {round(mean_val, 2)})"
                        })
    
    # Generate smart recommendations
    total_trends = len([t for t in insights["data_trends"] if t["direction"] != "stabil"])
    positive_trends = len([t for t in insights["data_trends"] if t["direction"] == "artış"])
    negative_trends = len([t for t in insights["data_trends"] if t["direction"] == "azalış"])
    
    if positive_trends > negative_trends:
        insights["recommendations"].append("🎯 Genel performans pozitif yönde. Mevcut stratejileri sürdürün.")
        insights["performance_score"] = min(95, 70 + (positive_trends * 5))
    elif negative_trends > positive_trends:
        insights["recommendations"].append("⚠️ Bazı metriklerde düşüş gözleniyor. Aksiyon planı oluşturun.")
        insights["performance_score"] = max(30, 70 - (negative_trends * 8))
    else:
        insights["recommendations"].append("📊 Metrikler stabil durumda. Sürekli iyileştirme fırsatları araştırın.")
        insights["performance_score"] = 65
    
    if len(insights["anomalies"]) > 0:
        insights["recommendations"].append(f"🔍 {len(insights['anomalies'])} adet anormal veri tespit edildi. Detaylı inceleme önerilir.")
    
    if len(responses) >= 6:
        insights["confidence_level"] = "high"
        insights["recommendations"].append("✅ Yeterli veri mevcut. Analiz sonuçları güvenilir.")
    elif len(responses) >= 3:
        insights["confidence_level"] = "medium"
        insights["recommendations"].append("📈 Orta düzey veri mevcut. Daha fazla veri toplayarak analiz kalitesini artırabilirsiniz.")
    else:
        insights["confidence_level"] = "low"
        insights["recommendations"].append("📊 Sınırlı veri mevcut. Güvenilir trend analizi için daha fazla veri gerekli.")
    
    return {
        "question_id": question_id,
        "question_text": question.get("question_text", ""),
        "insights": insights,
        "data_points": len(responses),
        "periods_analyzed": len(period_data),
        "generated_at": datetime.now().isoformat()
    }

@api_router.get("/analytics/compare")
async def get_comparative_analytics(
    question_ids: str = Query(..., description="Comma-separated question IDs to compare"),
    current_user: User = Depends(get_current_user)
):
    """Compare analytics across multiple questions"""
    from datetime import datetime
    import statistics
    
    question_id_list = [qid.strip() for qid in question_ids.split(',') if qid.strip()]
    
    if len(question_id_list) < 2:
        raise HTTPException(status_code=400, detail="At least 2 question IDs required for comparison")
    
    if len(question_id_list) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 questions can be compared at once")
    
    comparison_results = []
    
    for question_id in question_id_list:
        # Get question data
        question = await db.questions.find_one({"id": question_id})
        if not question:
            continue
        
        # Get responses
        responses = await db.table_responses.find({"question_id": question_id}).to_list(1000)
        
        # Calculate basic metrics
        total_responses = len(responses)
        latest_response = max(responses, key=lambda x: f"{x.get('year', 0)}-{x.get('month', 0):02d}") if responses else None
        
        # Calculate trend if we have enough data
        trend_data = {}
        if len(responses) >= 2:
            table_rows = question.get("table_rows", [])
            for row in table_rows:
                values = []
                for response in sorted(responses, key=lambda x: f"{x.get('year', 0)}-{x.get('month', 0):02d}"):
                    value = response.get("table_data", {}).get(row["id"], "0")
                    try:
                        numeric_value = float(value) if value else 0
                        values.append(numeric_value)
                    except (ValueError, TypeError):
                        values.append(0)
                
                if len(values) >= 2:
                    recent_avg = statistics.mean(values[-3:]) if len(values) >= 3 else values[-1]
                    older_avg = statistics.mean(values[:-3]) if len(values) > 3 else values[0]
                    trend_percentage = ((recent_avg - older_avg) / older_avg * 100) if older_avg != 0 else 0
                    
                    trend_data[row["name"]] = {
                        "trend_percentage": round(trend_percentage, 2),
                        "direction": "artış" if trend_percentage > 5 else "azalış" if trend_percentage < -5 else "stabil",
                        "current_value": values[-1] if values else 0,
                        "average_value": round(statistics.mean(values), 2) if values else 0
                    }
        
        comparison_results.append({
            "question_id": question_id,
            "question_text": question.get("question_text", ""),
            "category": question.get("category", ""),
            "period": question.get("period", ""),
            "total_responses": total_responses,
            "latest_period": f"{latest_response.get('year', 0)}-{latest_response.get('month', 0):02d}" if latest_response else None,
            "trend_data": trend_data,
            "data_quality": "high" if total_responses >= 6 else "medium" if total_responses >= 3 else "low"
        })
    
    # Generate comparison insights
    insights = {
        "summary": {
            "questions_compared": len(comparison_results),
            "total_data_points": sum(r["total_responses"] for r in comparison_results),
            "high_quality_questions": len([r for r in comparison_results if r["data_quality"] == "high"]),
            "categories": list(set(r["category"] for r in comparison_results if r["category"]))
        },
        "performance_ranking": [],
        "trend_comparison": {},
        "recommendations": []
    }
    
    # Rank questions by data quality and trend performance
    for result in comparison_results:
        score = 0
        positive_trends = 0
        negative_trends = 0
        
        for metric, trend in result["trend_data"].items():
            if trend["direction"] == "artış":
                positive_trends += 1
                score += 10
            elif trend["direction"] == "azalış":
                negative_trends += 1
                score -= 5
            else:
                score += 2
        
        # Bonus for data quality
        if result["data_quality"] == "high":
            score += 20
        elif result["data_quality"] == "medium":
            score += 10
        
        insights["performance_ranking"].append({
            "question_id": result["question_id"],
            "question_text": result["question_text"],
            "score": score,
            "positive_trends": positive_trends,
            "negative_trends": negative_trends,
            "data_quality": result["data_quality"]
        })
    
    # Sort by performance score
    insights["performance_ranking"].sort(key=lambda x: x["score"], reverse=True)
    
    # Generate recommendations
    best_performer = insights["performance_ranking"][0] if insights["performance_ranking"] else None
    worst_performer = insights["performance_ranking"][-1] if insights["performance_ranking"] else None
    
    if best_performer and worst_performer and len(insights["performance_ranking"]) > 1:
        insights["recommendations"].append(f"🏆 En iyi performans: '{best_performer['question_text'][:50]}...' (Skor: {best_performer['score']})")
        insights["recommendations"].append(f"⚠️ Dikkat gereken: '{worst_performer['question_text'][:50]}...' (Skor: {worst_performer['score']})")
    
    high_quality_count = insights["summary"]["high_quality_questions"]
    total_questions = insights["summary"]["questions_compared"]
    
    if high_quality_count == total_questions:
        insights["recommendations"].append("✅ Tüm sorular yeterli veri kalitesine sahip.")
    elif high_quality_count > total_questions / 2:
        insights["recommendations"].append("📊 Çoğu soru iyi veri kalitesine sahip, bazıları daha fazla veri gerektirebilir.")
    else:
        insights["recommendations"].append("📈 Veri kalitesini artırmak için daha fazla yanıt toplanması önerilir.")
    
    return {
        "comparison_results": comparison_results,
        "insights": insights,
        "generated_at": datetime.now().isoformat()
    }

@api_router.post("/automation/email-reminders")
async def setup_email_reminders(
    reminder_config: dict,
    current_user: User = Depends(get_current_user)
):
    """Setup automated email reminders for pending responses"""
    from datetime import datetime, timedelta
    
    # Get all assignments that need reminders
    current_time = datetime.now()
    
    # Find assignments older than specified days without response
    reminder_days = reminder_config.get("reminder_days", 3)
    cutoff_date = current_time - timedelta(days=reminder_days)
    
    # Get assignments that need reminders
    assignments = await db.question_assignments.find({
        "sent_date": {"$lt": cutoff_date},
        "response_received": {"$ne": True}
    }).to_list(1000)
    
    reminder_count = 0
    reminders_sent = []
    
    for assignment in assignments:
        # Get question and employee details
        question = await db.questions.find_one({"id": assignment["question_id"]})
        employee = await db.employees.find_one({"id": assignment["employee_id"]})
        
        if not question or not employee:
            continue
            
        # Check if reminder already sent recently
        last_reminder = assignment.get("last_reminder_sent")
        if last_reminder:
            days_since_reminder = (current_time - datetime.fromisoformat(last_reminder)).days
            if days_since_reminder < reminder_config.get("min_reminder_interval", 2):
                continue
        
        # Prepare reminder email content
        reminder_data = {
            "type": "reminder",
            "employee_name": f"{employee['first_name']} {employee['last_name']}",
            "employee_email": employee["email"],
            "question_text": question["question_text"],
            "category": question["category"],
            "days_pending": (current_time - datetime.fromisoformat(assignment["sent_date"])).days,
            "response_url": f"http://localhost:3000/answer/{assignment['assignment_id']}",
            "assignment_id": assignment["assignment_id"]
        }
        
        # Send reminder email (demo)
        print(f"[REMINDER EMAIL] Hatırlatma e-postası gönderildi:")
        print(f"Alıcı: {employee['email']}")
        print(f"Konu: Hatırlatma - {question['category']} Sorusu Yanıt Bekliyor")
        print(f"Gecikme: {reminder_data['days_pending']} gün")
        print("---")
        
        # Update assignment with reminder info
        await db.question_assignments.update_one(
            {"assignment_id": assignment["assignment_id"]},
            {
                "$set": {
                    "last_reminder_sent": current_time.isoformat(),
                    "reminder_count": assignment.get("reminder_count", 0) + 1
                }
            }
        )
        
        reminders_sent.append(reminder_data)
        reminder_count += 1
    
    return {
        "success": True,
        "reminder_count": reminder_count,
        "reminders_sent": reminders_sent,
        "config": reminder_config,
        "processed_at": current_time.isoformat()
    }

@api_router.post("/automation/generate-reports")
async def generate_automated_reports(
    report_config: dict,
    current_user: User = Depends(get_current_user)
):
    """Generate automated reports based on configuration"""
    from datetime import datetime, timedelta
    import json
    
    current_time = datetime.now()
    report_type = report_config.get("type", "monthly")
    
    # Calculate report period
    if report_type == "weekly":
        start_date = current_time - timedelta(days=7)
        period_name = "Haftalık"
    elif report_type == "monthly":
        start_date = current_time.replace(day=1)
        period_name = "Aylık"
    elif report_type == "quarterly":
        # Get start of quarter
        quarter_start_month = ((current_time.month - 1) // 3) * 3 + 1
        start_date = current_time.replace(month=quarter_start_month, day=1)
        period_name = "Çeyreklik"
    else:
        start_date = current_time - timedelta(days=30)
        period_name = "Özel"
    
    # Get responses in the period
    responses = await db.table_responses.find({
        "created_at": {"$gte": start_date}
    }).to_list(1000)
    
    # Get all questions for context
    questions = await db.questions.find({}).to_list(1000)
    question_map = {q["id"]: q for q in questions}
    
    # Generate report data
    report_data = {
        "period": period_name,
        "start_date": start_date.isoformat(),
        "end_date": current_time.isoformat(),
        "total_responses": len(responses),
        "questions_analyzed": len(set(r["question_id"] for r in responses)),
        "response_breakdown": {},
        "top_performers": [],
        "insights": [],
        "recommendations": []
    }
    
    # Analyze responses by question
    question_stats = {}
    for response in responses:
        question_id = response["question_id"]
        if question_id not in question_stats:
            question_stats[question_id] = {
                "question_text": question_map.get(question_id, {}).get("question_text", "Unknown"),
                "category": question_map.get(question_id, {}).get("category", "Unknown"),
                "response_count": 0,
                "total_values": []
            }
        
        question_stats[question_id]["response_count"] += 1
        
        # Extract numeric values from table_data
        table_data = response.get("table_data", {})
        for value in table_data.values():
            try:
                numeric_value = float(value) if value else 0
                question_stats[question_id]["total_values"].append(numeric_value)
            except (ValueError, TypeError):
                pass
    
    # Generate insights
    if len(responses) > 0:
        report_data["insights"].append(f"📊 {period_name} dönemde {len(responses)} yanıt alındı")
        report_data["insights"].append(f"📈 {len(question_stats)} farklı soru kategorisi analiz edildi")
        
        # Find most active category
        category_counts = {}
        for stats in question_stats.values():
            category = stats["category"]
            category_counts[category] = category_counts.get(category, 0) + stats["response_count"]
        
        if category_counts:
            top_category = max(category_counts.items(), key=lambda x: x[1])
            report_data["insights"].append(f"🏆 En aktif kategori: {top_category[0]} ({top_category[1]} yanıt)")
    
    # Generate recommendations
    if len(responses) < 5:
        report_data["recommendations"].append("📈 Yanıt oranını artırmak için hatırlatma e-postaları göndermeyi düşünün")
    else:
        report_data["recommendations"].append("✅ İyi performans gösteriyorsunuz, bu tempoyu koruyun")
    
    if len(question_stats) < 3:
        report_data["recommendations"].append("📊 Daha kapsamlı analiz için farklı kategorilerde sorular eklemeyi düşünün")
    
    # Store report in database for future reference
    report_record = {
        "id": f"report_{current_time.strftime('%Y%m%d_%H%M%S')}",
        "type": report_type,
        "generated_at": current_time.isoformat(),
        "generated_by": current_user.username if current_user else "system",
        "data": report_data,
        "config": report_config
    }
    
    await db.automated_reports.insert_one(report_record)
    
    # Demo: Print report summary
    print(f"[AUTOMATED REPORT] {period_name} Raporu Oluşturuldu:")
    print(f"Dönem: {start_date.strftime('%Y-%m-%d')} - {current_time.strftime('%Y-%m-%d')}")
    print(f"Toplam Yanıt: {len(responses)}")
    print(f"Analiz Edilen Sorular: {len(question_stats)}")
    for insight in report_data["insights"]:
        print(f"• {insight}")
    print("---")
    
    return {
        "success": True,
        "report_id": report_record["id"],
        "report_data": report_data,
        "generated_at": current_time.isoformat()
    }

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    """Get dynamic dashboard statistics"""
    from datetime import datetime, timedelta
    
    current_time = datetime.now()
    current_month = current_time.month
    current_year = current_time.year
    
    # Calculate previous month for trend calculation
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year
    
    # Get current month responses
    current_month_responses = await db.table_responses.count_documents({
        "year": current_year,
        "month": current_month
    })
    
    # Get previous month responses for trend calculation
    prev_month_responses = await db.table_responses.count_documents({
        "year": prev_year,
        "month": prev_month
    })
    
    # Calculate monthly trend
    if prev_month_responses > 0:
        monthly_trend = ((current_month_responses - prev_month_responses) / prev_month_responses) * 100
    else:
        monthly_trend = 0 if current_month_responses == 0 else 100
    
    # Get total active users (employees)
    active_users = await db.employees.count_documents({})
    
    # Get total active questions
    active_questions = await db.questions.count_documents({})
    
    # Get responses with AI comments
    ai_analyses = await db.table_responses.count_documents({
        "ai_comment": {"$exists": True, "$ne": None, "$ne": ""}
    })
    
    # Calculate completion rate (responses vs expected responses)
    # For simplicity, assume each active question should have 1 response per employee per month
    expected_responses = active_questions * active_users
    completion_rate = (current_month_responses / expected_responses * 100) if expected_responses > 0 else 0
    
    # Generate dynamic notifications
    notifications = []
    
    # Notification about pending responses
    pending_responses = max(0, expected_responses - current_month_responses)
    if pending_responses > 0:
        notifications.append({
            "type": "warning",
            "message": f"{pending_responses} soru yanıt bekliyor",
            "priority": "medium"
        })
    
    # Notification about AI analyses
    if ai_analyses > 0:
        notifications.append({
            "type": "info", 
            "message": f"{ai_analyses} AI analizi hazır",
            "priority": "low"
        })
    else:
        notifications.append({
            "type": "info",
            "message": "0 AI analizi hazır",
            "priority": "low"
        })
    
    # Monthly report notification
    notifications.append({
        "type": "info",
        "message": "Aylık rapor hazırlanıyor",
        "priority": "medium"
    })
    
    return {
        "monthly_responses": current_month_responses,
        "monthly_trend": round(monthly_trend, 1),
        "active_users": active_users,
        "completion_rate": round(completion_rate, 1),
        "ai_analyses": ai_analyses,
        "active_questions": active_questions,
        "notifications": notifications,
        "last_updated": current_time.isoformat()
    }

# Original routes (now protected)
@api_router.get("/")
async def root():
    return {"message": "Auth System API"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate, current_user: User = Depends(get_current_user)):
    status_dict = input.dict()
    status_obj = StatusCheck(**status_dict)
    status_data = status_obj.dict()
    status_data["created_at"] = status_data["timestamp"].isoformat()
    status_data.pop("timestamp", None)
    _ = await db.status_checks.insert_one(status_data)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks(current_user: User = Depends(get_current_user)):
    status_checks = await db.status_checks.find().to_list(1000)
    result = []
    for status_check in status_checks:
        if "created_at" in status_check:
            status_check["timestamp"] = datetime.fromisoformat(status_check["created_at"].replace('Z', '+00:00')) if isinstance(status_check["created_at"], str) else status_check["created_at"]
        result.append(StatusCheck(**status_check))
    return result

# Employee Routes
@api_router.post("/employees", response_model=Employee)
async def create_employee(employee_data: EmployeeCreate, current_user: User = Depends(get_current_user)):
    # Validate date formats
    try:
        datetime.strptime(employee_data.hire_date, "%Y-%m-%d")
        datetime.strptime(employee_data.birth_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tarih formatı YYYY-MM-DD olmalıdır"
        )
    
    # Check if employee with same phone already exists
    existing_employee = await db.employees.find_one({"phone": employee_data.phone})
    if existing_employee:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu telefon numarası ile kayıtlı çalışan zaten mevcut"
        )
    
    employee_dict = employee_data.dict()
    employee_dict["id"] = str(uuid.uuid4())
    employee_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.employees.insert_one(employee_dict)
    
    employee = Employee(**employee_dict)
    employee.created_at = datetime.fromisoformat(employee_dict["created_at"].replace('Z', '+00:00')) if isinstance(employee_dict["created_at"], str) else employee_dict["created_at"]
    
    return employee

@api_router.get("/employees", response_model=List[Employee])
async def get_employees(current_user: User = Depends(get_current_user)):
    employees = await db.employees.find().to_list(1000)
    result = []
    for employee in employees:
        if "created_at" in employee:
            employee["created_at"] = datetime.fromisoformat(employee["created_at"].replace('Z', '+00:00')) if isinstance(employee["created_at"], str) else employee["created_at"]
        result.append(Employee(**employee))
    return result

@api_router.get("/employees/{employee_id}", response_model=Employee)
async def get_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    employee = await db.employees.find_one({"id": employee_id})
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Çalışan bulunamadı"
        )
    
    if "created_at" in employee:
        employee["created_at"] = datetime.fromisoformat(employee["created_at"].replace('Z', '+00:00')) if isinstance(employee["created_at"], str) else employee["created_at"]
    
    return Employee(**employee)

@api_router.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: str, employee_data: EmployeeCreate, current_user: User = Depends(get_current_user)):
    # Validate date formats
    try:
        datetime.strptime(employee_data.hire_date, "%Y-%m-%d")
        datetime.strptime(employee_data.birth_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tarih formatı YYYY-MM-DD olmalıdır"
        )
    
    # Check if employee exists
    existing_employee = await db.employees.find_one({"id": employee_id})
    if not existing_employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Çalışan bulunamadı"
        )
    
    # Check if phone is taken by another employee
    phone_check = await db.employees.find_one({"phone": employee_data.phone, "id": {"$ne": employee_id}})
    if phone_check:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu telefon numarası ile kayıtlı başka bir çalışan mevcut"
        )
    
    update_data = employee_data.dict()
    await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    updated_employee = await db.employees.find_one({"id": employee_id})
    if "created_at" in updated_employee:
        updated_employee["created_at"] = datetime.fromisoformat(updated_employee["created_at"].replace('Z', '+00:00')) if isinstance(updated_employee["created_at"], str) else updated_employee["created_at"]
    
    return Employee(**updated_employee)

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    result = await db.employees.delete_one({"id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Çalışan bulunamadı"
        )
    
    return {"message": "Çalışan başarıyla silindi"}

# Question Bank Routes
@api_router.post("/questions", response_model=Question)
async def create_question(question_data: QuestionCreate, current_user: User = Depends(get_current_user)):
    question_dict = question_data.dict()
    question_dict["id"] = str(uuid.uuid4())
    question_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.questions.insert_one(question_dict)
    
    question = Question(**question_dict)
    question.created_at = datetime.fromisoformat(question_dict["created_at"].replace('Z', '+00:00')) if isinstance(question_dict["created_at"], str) else question_dict["created_at"]
    
    return question

@api_router.get("/questions", response_model=List[Question])
async def get_questions(current_user: User = Depends(get_current_user)):
    questions = await db.questions.find().to_list(1000)
    result = []
    for question in questions:
        if "created_at" in question:
            question["created_at"] = datetime.fromisoformat(question["created_at"].replace('Z', '+00:00')) if isinstance(question["created_at"], str) else question["created_at"]
        result.append(Question(**question))
    return result

@api_router.get("/questions/{question_id}", response_model=Question)
async def get_question(question_id: str, current_user: User = Depends(get_current_user)):
    question = await db.questions.find_one({"id": question_id})
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru bulunamadı"
        )
    
    if "created_at" in question:
        question["created_at"] = datetime.fromisoformat(question["created_at"].replace('Z', '+00:00')) if isinstance(question["created_at"], str) else question["created_at"]
    
    return Question(**question)

@api_router.put("/questions/{question_id}", response_model=Question)
async def update_question(question_id: str, question_data: QuestionCreate, current_user: User = Depends(get_current_user)):
    # Check if question exists
    existing_question = await db.questions.find_one({"id": question_id})
    if not existing_question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru bulunamadı"
        )
    
    update_data = question_data.dict()
    await db.questions.update_one({"id": question_id}, {"$set": update_data})
    
    updated_question = await db.questions.find_one({"id": question_id})
    if "created_at" in updated_question:
        updated_question["created_at"] = datetime.fromisoformat(updated_question["created_at"].replace('Z', '+00:00')) if isinstance(updated_question["created_at"], str) else updated_question["created_at"]
    
    return Question(**updated_question)

@api_router.delete("/questions/{question_id}")
async def delete_question(question_id: str, current_user: User = Depends(get_current_user)):
    result = await db.questions.delete_one({"id": question_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru bulunamadı"
        )
    
    return {"message": "Soru başarıyla silindi"}

# Category Routes
@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, current_user: User = Depends(get_current_user)):
    # Check if category already exists
    existing_category = await db.categories.find_one({"name": category_data.name})
    if existing_category:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu kategori zaten mevcut"
        )
    
    category_dict = category_data.dict()
    category_dict["id"] = str(uuid.uuid4())
    category_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.categories.insert_one(category_dict)
    
    category = Category(**category_dict)
    category.created_at = datetime.fromisoformat(category_dict["created_at"].replace('Z', '+00:00')) if isinstance(category_dict["created_at"], str) else category_dict["created_at"]
    
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories(current_user: User = Depends(get_current_user)):
    categories = await db.categories.find().sort("name", 1).to_list(1000)
    result = []
    for category in categories:
        if "created_at" in category:
            category["created_at"] = datetime.fromisoformat(category["created_at"].replace('Z', '+00:00')) if isinstance(category["created_at"], str) else category["created_at"]
        result.append(Category(**category))
    return result

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: User = Depends(get_current_user)):
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Kategori bulunamadı"
        )
    
    return {"message": "Kategori başarıyla silindi"}

# Department Routes
@api_router.post("/departments", response_model=Department)
async def create_department(department_data: DepartmentCreate, current_user: User = Depends(get_current_user)):
    # Check if department already exists
    existing_department = await db.departments.find_one({"name": department_data.name})
    if existing_department:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu departman zaten mevcut"
        )
    
    department_dict = department_data.dict()
    department_dict["id"] = str(uuid.uuid4())
    department_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.departments.insert_one(department_dict)
    
    department = Department(**department_dict)
    department.created_at = datetime.fromisoformat(department_dict["created_at"].replace('Z', '+00:00')) if isinstance(department_dict["created_at"], str) else department_dict["created_at"]
    
    return department

@api_router.get("/departments", response_model=List[Department])
async def get_departments(current_user: User = Depends(get_current_user)):
    departments = await db.departments.find().sort("name", 1).to_list(1000)
    result = []
    for department in departments:
        if "created_at" in department:
            department["created_at"] = datetime.fromisoformat(department["created_at"].replace('Z', '+00:00')) if isinstance(department["created_at"], str) else department["created_at"]
        result.append(Department(**department))
    return result

@api_router.delete("/departments/{department_id}")
async def delete_department(department_id: str, current_user: User = Depends(get_current_user)):
    result = await db.departments.delete_one({"id": department_id})
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Departman bulunamadı"
        )
    
    return {"message": "Departman başarıyla silindi"}

# Question Sharing Routes
@api_router.get("/questions-share-list")
async def get_questions_for_sharing(
    period: Optional[str] = Query(None, description="Filter questions by period"),
    current_user: User = Depends(get_current_user)
):
    """Get all questions with employees for sharing interface"""
    # Build filter query
    filter_query = {}
    if period:
        # Validate period value
        valid_periods = ["Günlük", "Haftalık", "Aylık", "Çeyreklik", "Altı Aylık", "Yıllık", "İhtiyaç Halinde"]
        if period in valid_periods:
            filter_query["period"] = period
    
    # Get questions with optional filtering
    questions = await db.questions.find(filter_query).to_list(1000)
    employees = await db.employees.find().to_list(1000)
    
    # Format questions
    formatted_questions = []
    for question in questions:
        if "created_at" in question:
            question["created_at"] = datetime.fromisoformat(question["created_at"].replace('Z', '+00:00')) if isinstance(question["created_at"], str) else question["created_at"]
        formatted_questions.append(Question(**question))
    
    # Format employees
    formatted_employees = []
    for employee in employees:
        if "created_at" in employee:
            employee["created_at"] = datetime.fromisoformat(employee["created_at"].replace('Z', '+00:00')) if isinstance(employee["created_at"], str) else employee["created_at"]
        formatted_employees.append(Employee(**employee))
    
    return {
        "questions": formatted_questions,
        "employees": formatted_employees
    }

@api_router.post("/questions-share")
async def share_questions(share_request: ShareQuestionsRequest, current_user: User = Depends(get_current_user)):
    """Share questions via email to assigned employees"""
    current_date = datetime.now(timezone.utc)
    year = current_date.year
    month = current_date.month
    month_names = {
        1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
        7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
    }
    month_year = f"{month_names[month]} {year}"
    
    assignments_created = []
    email_successes = 0
    email_failures = []
    resend_count = 0
    
    for assignment_data in share_request.assignments:
        question_id = assignment_data.get("question_id")
        employee_id = assignment_data.get("employee_id")
        
        if not question_id or not employee_id:
            continue
        
        # Check if assignment already exists for this month  
        existing_assignment = await db.question_assignments.find_one({
            "question_id": question_id,
            "employee_id": employee_id, 
            "year": year,
            "month": month
        })
        
        # Allow re-sending: Don't skip if already assigned, just log it
        is_resend = bool(existing_assignment)
        if is_resend:
            resend_count += 1
        
        # Get question and employee details
        question = await db.questions.find_one({"id": question_id})
        employee = await db.employees.find_one({"id": employee_id})
        
        if not question or not employee:
            continue
        
        # Create or update assignment  
        if existing_assignment:
            # Update existing assignment for re-send
            assignment_id = existing_assignment["id"]
            assignment_dict = existing_assignment.copy()
            assignment_dict["assigned_at"] = current_date.isoformat()  # Update timestamp
            assignment_dict["email_sent"] = False  # Reset email status
        else:
            # Create new assignment
            assignment_id = str(uuid.uuid4())
            assignment_dict = {
                "id": assignment_id,
                "question_id": question_id,
                "employee_id": employee_id,
                "year": year,
                "month": month,
                "email_sent": False,
                "response_received": False,
                "assigned_at": current_date.isoformat()
            }
        
        # Send email if employee has email address
        email_sent = False
        if employee.get('email'):
            employee_name = f"{employee['first_name']} {employee['last_name']}"
            email_sent = await send_question_email(
                employee['email'],
                employee_name,
                question,
                assignment_id,
                month_year
            )
            
            if email_sent:
                email_successes += 1
            else:
                email_failures.append(employee_name)
        else:
            # No email address
            email_failures.append(f"{employee['first_name']} {employee['last_name']} (E-posta adresi yok)")
        
        # Update assignment with email status
        assignment_dict["email_sent"] = email_sent
        
        if existing_assignment:
            # Update existing assignment
            await db.question_assignments.update_one(
                {"id": assignment_id},
                {"$set": assignment_dict}
            )
        else:
            # Insert new assignment
            await db.question_assignments.insert_one(assignment_dict)
        
        assignments_created.append(assignment_dict)
    
    # Prepare response message
    message_parts = []
    if assignments_created:
        new_assignments = len(assignments_created) - resend_count
        if new_assignments > 0:
            message_parts.append(f"{new_assignments} yeni soru atandı")
        if resend_count > 0:
            message_parts.append(f"{resend_count} soru tekrar gönderildi")
    
    if email_successes:
        message_parts.append(f"{email_successes} e-posta başarıyla gönderildi")
    
    if email_failures:
        message_parts.append(f"{len(email_failures)} e-posta gönderilemedi")
    
    response_message = ", ".join(message_parts) if message_parts else "İşlem tamamlandı"
    
    return {
        "message": response_message,
        "assignments_created": len(assignments_created),
        "emails_sent": email_successes,
        "email_failures": email_failures,
        "year": year,
        "month": month
    }

# Public Question Response Routes (No auth required)
@api_router.get("/public/question-form/{assignment_id}")
async def get_public_question_form(assignment_id: str):
    """Get question form for public response (no auth required)"""
    assignment = await db.question_assignments.find_one({"id": assignment_id})
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru ataması bulunamadı"
        )
    
    # Check if already responded
    existing_response = await db.question_responses.find_one({"assignment_id": assignment_id})
    if existing_response:
        return {"message": "Bu soruya zaten yanıt verilmiş", "already_responded": True}
    
    # Get question and employee details
    question = await db.questions.find_one({"id": assignment["question_id"]})
    employee = await db.employees.find_one({"id": assignment["employee_id"]})
    
    if not question or not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru veya çalışan bulunamadı"
        )
    
    return {
        "assignment_id": assignment_id,
        "question": {
            "id": question["id"],
            "question_text": question["question_text"],
            "category": question["category"],
            "importance_reason": question["importance_reason"],
            "expected_action": question["expected_action"],
            "period": question["period"],  # ✅ Period bilgisi eklendi
            "table_rows": question.get("table_rows", [])  # ✅ Table rows eklendi
        },
        "employee": {
            "first_name": employee["first_name"],
            "last_name": employee["last_name"],
            "department": employee["department"]
        },
        "year": assignment["year"],
        "month": assignment["month"],
        "already_responded": False
    }

@api_router.post("/public/question-response")
async def submit_public_question_response(response_data: QuestionResponseCreate):
    """Submit question response from public form (no auth required)"""
    # Verify assignment exists and not already responded
    assignment = await db.question_assignments.find_one({"id": response_data.assignment_id})
    if not assignment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru ataması bulunamadı"
        )
    
    # Check if already responded
    existing_response = await db.question_responses.find_one({"assignment_id": response_data.assignment_id})
    if existing_response:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bu soruya zaten yanıt verilmiş"
        )
    
    # Create response
    response_dict = {
        "id": str(uuid.uuid4()),
        "assignment_id": response_data.assignment_id,
        "question_id": assignment["question_id"],
        "employee_id": assignment["employee_id"],
        "response_text": response_data.response_text,
        "year": assignment["year"],
        "month": assignment["month"],
        "submitted_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.question_responses.insert_one(response_dict)
    
    # Update assignment as responded
    await db.question_assignments.update_one(
        {"id": response_data.assignment_id},
        {"$set": {"response_received": True}}
    )
    
    return {
        "message": "Yanıtınız başarıyla kaydedildi",
        "response_id": response_dict["id"]
    }

# Question Responses Management
@api_router.get("/question-responses")
async def get_question_responses(current_user: User = Depends(get_current_user)):
    """Get all question responses"""
    responses = await db.question_responses.find().to_list(1000)
    
    # Get additional data for each response
    formatted_responses = []
    for response in responses:
        question = await db.questions.find_one({"id": response["question_id"]})
        employee = await db.employees.find_one({"id": response["employee_id"]})
        
        if question and employee:
            formatted_response = {
                **response,
                "question_text": question["question_text"],
                "question_category": question["category"],
                "employee_name": f"{employee['first_name']} {employee['last_name']}",
                "employee_department": employee["department"]
            }
            formatted_responses.append(formatted_response)
    
    return formatted_responses

@api_router.get("/email-logs")
async def get_email_logs(current_user: User = Depends(get_current_user)):
    """Get demo email logs"""
    assignments = await db.question_assignments.find({"email_sent": True}).sort("assigned_at", -1).to_list(20)
    
    email_logs = []
    for assignment in assignments:
        question = await db.questions.find_one({"id": assignment["question_id"]})
        employee = await db.employees.find_one({"id": assignment["employee_id"]})
        
        if question and employee:
            email_log = {
                "assignment_id": assignment["id"],
                "employee_name": f"{employee['first_name']} {employee['last_name']}",
                "employee_email": employee.get('email', 'E-posta yok'),
                "question_category": question["category"],
                "question_text": question["question_text"],
                "answer_link": f"{frontend_url}/answer/{assignment['id']}",
                "sent_date": assignment["assigned_at"],
                "response_received": assignment.get("response_received", False),
                "year": assignment["year"],
                "month": assignment["month"]
            }
            email_logs.append(email_log)
    
    return email_logs

@api_router.get("/answer-status")
async def get_answer_status(current_user: User = Depends(get_current_user)):
    """Get all question assignments with response status"""
    assignments = await db.question_assignments.find().sort("assigned_at", -1).to_list(1000)
    
    answer_status_list = []
    for assignment in assignments:
        question = await db.questions.find_one({"id": assignment["question_id"]})
        employee = await db.employees.find_one({"id": assignment["employee_id"]})
        
        # Get response if exists
        response = await db.question_responses.find_one({"assignment_id": assignment["id"]})
        
        if question and employee:
            status_item = {
                "assignment_id": assignment["id"],
                "question": {
                    "id": question["id"],
                    "category": question["category"],
                    "question_text": question["question_text"],
                    "importance_reason": question.get("importance_reason", ""),
                    "expected_action": question.get("expected_action", "")
                },
                "employee": {
                    "id": employee["id"],
                    "name": f"{employee['first_name']} {employee['last_name']}",
                    "email": employee.get('email', 'E-posta yok'),
                    "department": employee["department"]
                },
                "assignment_date": assignment["assigned_at"],
                "year": assignment["year"],
                "month": assignment["month"],
                "email_sent": assignment.get("email_sent", False),
                "response_received": assignment.get("response_received", False),
                "response": {
                    "submitted": bool(response),
                    "text": response.get("response_text", "") if response else "",
                    "submitted_at": response.get("submitted_at", "") if response else ""
                }
            }
            answer_status_list.append(status_item)
    
    return answer_status_list

# Table Response Management - Clean System
@api_router.get("/table-responses")
async def get_table_responses(current_user: User = Depends(get_current_user)):
    """Get all table responses with question and employee details"""
    responses = await db.table_responses.find().to_list(1000)
    
    formatted_responses = []
    for response in responses:
        response.pop('_id', None)
        
        question = await db.questions.find_one({"id": response["question_id"]})
        employee = await db.employees.find_one({"id": response["employee_id"]})
        
        if question and employee:
            question.pop('_id', None)
            employee.pop('_id', None)
            
            formatted_response = {
                **response,
                "question": {
                    "id": question["id"],
                    "question_text": question["question_text"],
                    "category": question["category"],
                    "period": question["period"],
                    "table_rows": question.get("table_rows", [])
                },
                "employee": {
                    "id": employee["id"],
                    "name": f"{employee['first_name']} {employee['last_name']}",
                    "department": employee["department"]
                }
            }
            formatted_responses.append(formatted_response)
    
    return formatted_responses

@api_router.get("/table-responses/question/{question_id}")
async def get_responses_by_question(question_id: str, current_user: User = Depends(get_current_user)):
    """Get all table responses for a specific question"""
    responses = await db.table_responses.find({"question_id": question_id}).to_list(1000)
    
    question = await db.questions.find_one({"id": question_id})
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru bulunamadı"
        )
    
    formatted_responses = []
    for response in responses:
        response.pop('_id', None)
        employee = await db.employees.find_one({"id": response["employee_id"]})
        if employee:
            employee.pop('_id', None)
            formatted_response = {
                **response,
                "employee": {
                    "id": employee["id"],
                    "name": f"{employee['first_name']} {employee['last_name']}",
                    "department": employee["department"]
                }
            }
            formatted_responses.append(formatted_response)
    
    question.pop('_id', None)
    return {
        "question": question,
        "responses": formatted_responses
    }

# Original table response endpoint removed - replaced with new implementation below

@api_router.get("/questions-for-responses") 
async def get_questions_for_responses(current_user: User = Depends(get_current_user)):
    """Get all questions with employee list for response management"""
    questions = await db.questions.find().to_list(1000)
    employees = await db.employees.find().to_list(1000)
    
    # Format questions
    formatted_questions = []
    for question in questions:
        if "created_at" in question:
            question["created_at"] = datetime.fromisoformat(question["created_at"].replace('Z', '+00:00')) if isinstance(question["created_at"], str) else question["created_at"]
        formatted_questions.append(Question(**question))
    
    # Format employees  
    formatted_employees = []
    for employee in employees:
        if "created_at" in employee:
            employee["created_at"] = datetime.fromisoformat(employee["created_at"].replace('Z', '+00:00')) if isinstance(employee["created_at"], str) else employee["created_at"]
        formatted_employees.append(Employee(**employee))
    
    return {
        "questions": formatted_questions,
        "employees": formatted_employees
    }

@api_router.post("/table-responses")
async def create_table_response(response_data: TableResponseCreate, current_user: User = Depends(get_current_user)):
    """Create or update a single table response"""
    try:
        # Check if question and employee exist
        question = await db.questions.find_one({"id": response_data.question_id})
        employee = await db.employees.find_one({"id": response_data.employee_id})
        
        if not question:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Soru bulunamadı"
            )
        
        if not employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Çalışan bulunamadı"
            )
        
        # Check if response already exists
        existing_response = await db.table_responses.find_one({
            "question_id": response_data.question_id,
            "employee_id": response_data.employee_id,
            "year": response_data.year,
            "month": response_data.month
        })
        
        # Generate AI comment
        ai_comment = None
        if response_data.table_data or response_data.monthly_comment:
            ai_comment = await generate_ai_comment(
                question_text=question["question_text"],
                category=question["category"],
                period=question["period"],
                table_data=response_data.table_data,
                table_rows=question.get("table_rows", []),
                monthly_comment=response_data.monthly_comment,
                year=response_data.year,
                month=response_data.month
            )
        
        current_time = datetime.now(timezone.utc)
        
        if existing_response:
            # Update existing response
            update_data = {
                "table_data": response_data.table_data,
                "monthly_comment": response_data.monthly_comment,
                "ai_comment": ai_comment,
                "updated_at": current_time.isoformat()
            }
            
            await db.table_responses.update_one(
                {"id": existing_response["id"]},
                {"$set": update_data}
            )
            
            return {"success": True, "message": "Cevap güncellendi", "action": "updated"}
        
        else:
            # Create new response
            response_dict = response_data.dict()
            response_dict["id"] = str(uuid.uuid4())
            response_dict["ai_comment"] = ai_comment
            response_dict["created_at"] = current_time.isoformat()
            response_dict["updated_at"] = current_time.isoformat()
            
            await db.table_responses.insert_one(response_dict)
            
            return {"success": True, "message": "Cevap kaydedildi", "action": "created"}
        
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Cevap kaydetme hatası: {str(e)}"
        )

@api_router.post("/table-responses/bulk")
async def create_bulk_table_responses(responses_data: List[TableResponseCreate], current_user: User = Depends(get_current_user)):
    """Create multiple table responses at once"""
    try:
        results = []
        for response_data in responses_data:
            # Skip empty responses
            has_data = (
                response_data.table_data or 
                (response_data.monthly_comment and response_data.monthly_comment.strip())
            )
            if not has_data:
                continue
            
            question = await db.questions.find_one({"id": response_data.question_id})
            employee = await db.employees.find_one({"id": response_data.employee_id})
            
            if not question or not employee:
                continue
            
            existing_response = await db.table_responses.find_one({
                "question_id": response_data.question_id,
                "employee_id": response_data.employee_id,
                "year": response_data.year,
                "month": response_data.month
            })
            
            # Generate AI comment
            ai_comment = None
            if response_data.table_data or (response_data.monthly_comment and response_data.monthly_comment.strip()):
                ai_comment = await generate_ai_comment(
                    question_text=question["question_text"],
                    category=question["category"],
                    period=question["period"],
                    table_data=response_data.table_data,
                    table_rows=question.get("table_rows", []),
                    monthly_comment=response_data.monthly_comment,
                    year=response_data.year,
                    month=response_data.month
                )
            
            current_time = datetime.now(timezone.utc)
            
            if existing_response:
                update_data = {
                    "table_data": response_data.table_data,
                    "monthly_comment": response_data.monthly_comment,
                    "ai_comment": ai_comment,
                    "updated_at": current_time.isoformat()
                }
                await db.table_responses.update_one(
                    {"id": existing_response["id"]},
                    {"$set": update_data}
                )
                results.append({"id": existing_response["id"], "action": "updated"})
            else:
                response_dict = response_data.dict()
                response_dict["id"] = str(uuid.uuid4())
                response_dict["ai_comment"] = ai_comment
                response_dict["created_at"] = current_time.isoformat()
                response_dict["updated_at"] = current_time.isoformat()
                
                await db.table_responses.insert_one(response_dict)
                results.append({"id": response_dict["id"], "action": "created"})
        
        return {
            "success": True,
            "message": f"{len(results)} cevap başarıyla kaydedildi",
            "results": results
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Toplu cevap kaydetme hatası: {str(e)}"
        )

@api_router.get("/table-responses/summary/{question_id}")
async def get_table_summary(question_id: str, current_user: User = Depends(get_current_user)):
    """Get table response summary for a specific question"""
    question = await db.questions.find_one({"id": question_id})
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Soru bulunamadı"
        )
    
    responses = await db.table_responses.find({"question_id": question_id}).to_list(1000)
    
    # Group responses by month
    monthly_data = {}
    for month in range(1, 13):
        month_responses = [r for r in responses if r["month"] == month]
        monthly_data[month] = {
            "count": len(month_responses),
            "responses": month_responses
        }
    
    months = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", 
              "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    summary_data = []
    for i, month_name in enumerate(months, 1):
        summary_data.append({
            "month": month_name,
            "month_number": i,
            "response_count": monthly_data[i]["count"],
            "responses": monthly_data[i]["responses"]
        })
    
    question.pop('_id', None)
    
    return {
        "question": question,
        "summary_data": summary_data,
        "total_responses": len(responses)
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Export endpoints
@app.get("/api/export/questions/excel")
async def export_questions_excel(current_user: dict = Depends(get_current_user)):
    """Export questions to Excel format"""
    try:
        questions = await db.questions.find({}).to_list(length=None)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorular"
        
        # Headers
        headers = ['ID', 'Kategori', 'Soru Metni', 'Periyod', 'Grafik Tipi', 'Önem Sebebi', 'Beklenen Aksiyon', 'Oluşturma Tarihi']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_num, question in enumerate(questions, 2):
            ws.cell(row=row_num, column=1, value=question['id'])
            ws.cell(row=row_num, column=2, value=question.get('category', ''))
            ws.cell(row=row_num, column=3, value=question.get('question_text', ''))
            ws.cell(row=row_num, column=4, value=question.get('period', ''))
            ws.cell(row=row_num, column=5, value=question.get('chart_type', ''))
            ws.cell(row=row_num, column=6, value=question.get('importance_reason', ''))
            ws.cell(row=row_num, column=7, value=question.get('expected_action', ''))
            ws.cell(row=row_num, column=8, value=question.get('created_at', ''))
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sorular.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail="Excel export failed")

@app.get("/api/export/employees/excel")
async def export_employees_excel(current_user: dict = Depends(get_current_user)):
    """Export employees to Excel format"""
    try:
        employees = await db.employees.find({}).to_list(length=None)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Çalışanlar"
        
        # Headers
        headers = ['ID', 'Ad Soyad', 'E-posta', 'Telefon', 'Departman', 'Pozisyon', 'Yaş', 'Cinsiyet', 'Maaş', 'İşe Başlama Tarihi']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_num, employee in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=employee['id'])
            ws.cell(row=row_num, column=2, value=employee.get('name', ''))
            ws.cell(row=row_num, column=3, value=employee.get('email', ''))
            ws.cell(row=row_num, column=4, value=employee.get('phone', ''))
            ws.cell(row=row_num, column=5, value=employee.get('department', ''))
            ws.cell(row=row_num, column=6, value=employee.get('position', ''))
            ws.cell(row=row_num, column=7, value=employee.get('age', ''))
            ws.cell(row=row_num, column=8, value=employee.get('gender', ''))
            ws.cell(row=row_num, column=9, value=employee.get('salary', ''))
            ws.cell(row=row_num, column=10, value=employee.get('hire_date', ''))
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=calisanlar.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail="Excel export failed")

@app.get("/api/export/responses/excel")
async def export_responses_excel(current_user: dict = Depends(get_current_user)):
    """Export responses to Excel format"""
    try:
        responses = await db.table_responses.find({}).to_list(length=None)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cevaplar"
        
        # Headers
        headers = ['ID', 'Soru ID', 'Çalışan ID', 'Yıl', 'Ay', 'Veri', 'Yorum', 'Oluşturma Tarihi']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_num, response in enumerate(responses, 2):
            ws.cell(row=row_num, column=1, value=response['id'])
            ws.cell(row=row_num, column=2, value=response.get('question_id', ''))
            ws.cell(row=row_num, column=3, value=response.get('employee_id', ''))
            ws.cell(row=row_num, column=4, value=response.get('year', ''))
            ws.cell(row=row_num, column=5, value=response.get('month', ''))
            # Convert data dict to string
            data_str = str(response.get('data', '')) if response.get('data') else ''
            ws.cell(row=row_num, column=6, value=data_str)
            ws.cell(row=row_num, column=7, value=response.get('comment', ''))
            ws.cell(row=row_num, column=8, value=response.get('created_at', ''))
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cevaplar.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail="Excel export failed")

@app.get("/api/export/questions/pdf")
async def export_questions_pdf(current_user: dict = Depends(get_current_user)):
    """Export questions to PDF format"""
    try:
        questions = await db.questions.find({}).to_list(length=None)
        
        # Create PDF buffer
        pdf_buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.darkblue,
            alignment=1
        )
        
        # Create story (content)
        story = []
        
        # Title
        story.append(Paragraph("Dijital Dönüşüm Sistemi - Sorular Raporu", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary
        summary_text = f"Toplam Soru Sayısı: {len(questions)}<br/>Rapor Tarihi: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Table data
        data = [['Kategori', 'Soru Metni', 'Periyod', 'Grafik Tipi']]
        
        for question in questions:
            row = [
                question.get('category', '')[:20] + '...' if len(question.get('category', '')) > 20 else question.get('category', ''),
                question.get('question_text', '')[:50] + '...' if len(question.get('question_text', '')) > 50 else question.get('question_text', ''),
                question.get('period', ''),
                question.get('chart_type', '')
            ]
            data.append(row)
        
        # Create table
        table = Table(data, colWidths=[1.5*inch, 3*inch, 1*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.getvalue()),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=sorular.pdf"}
        )
        
    except Exception as e:
        logger.error(f"PDF export error: {str(e)}")
        raise HTTPException(status_code=500, detail="PDF export failed")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()